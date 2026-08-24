import os
import re
import json
import random
import string
from datetime import datetime, timezone, timedelta
from typing import Optional

from fastapi import FastAPI, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker

DATABASE_URL = os.environ.get("DATABASE_URL", "")
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

engine = create_engine(DATABASE_URL, pool_pre_ping=True)
SessionLocal = sessionmaker(bind=engine)

app = FastAPI(title="Frosty - Pedidos de Secos API")

# Mapeamento: nome da aba na planilha de estoque -> id do sócio no banco
SOCIO_SHEET_CODES = {
    "socio_e_f_a_comercio_de_alimen": "EFA",
    "socio_e_m_h_d_comercio_de_alim": "EMHD",
    "socio_eff_comercio_de_alimento": "EFF",
    "socio_frosty_produtos_alimenti": "FROSTY",
}

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],   # troque pelo domínio do Netlify depois de publicar, se quiser restringir
    allow_methods=["*"],
    allow_headers=["*"],
)


def init_db():
    with engine.begin() as conn:
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS socios (
                id TEXT PRIMARY KEY,
                nome TEXT NOT NULL
            )
        """))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS lojas (
                id TEXT PRIMARY KEY,
                nome TEXT NOT NULL,
                cnpj TEXT,
                cidade TEXT,
                socio_id TEXT NOT NULL REFERENCES socios(id)
            )
        """))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS estoque (
                socio_id TEXT NOT NULL REFERENCES socios(id),
                codigo TEXT NOT NULL,
                descricao TEXT,
                und TEXT,
                qtd_emb TEXT,
                grupo TEXT,
                saldo NUMERIC NOT NULL DEFAULT 0,
                PRIMARY KEY (socio_id, codigo)
            )
        """))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS pedidos (
                id TEXT PRIMARY KEY,
                numero TEXT,
                loja_id TEXT NOT NULL,
                loja_nome TEXT,
                socio_id TEXT NOT NULL,
                data_hora TEXT NOT NULL,
                status TEXT NOT NULL,
                itens_json TEXT NOT NULL
            )
        """))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS config (
                key TEXT PRIMARY KEY,
                value TEXT
            )
        """))
        conn.execute(text("ALTER TABLE socios ADD COLUMN IF NOT EXISTS codigo TEXT"))
        conn.execute(text("ALTER TABLE estoque ADD COLUMN IF NOT EXISTS valor_acumulado NUMERIC DEFAULT 0"))
        conn.execute(text("ALTER TABLE estoque ADD COLUMN IF NOT EXISTS valor_unitario NUMERIC DEFAULT 0"))
        # mapeamento das abas da planilha de estoque (nome da aba -> sócio), usado na importação
        for socio_id, codigo in SOCIO_SHEET_CODES.items():
            conn.execute(text(
                "UPDATE socios SET codigo = :codigo WHERE id = :id"
            ), {"codigo": codigo, "id": socio_id})

        # ---- EPI / Material de Expediente ----
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS epi_setores (
                id TEXT PRIMARY KEY,
                nome TEXT NOT NULL,
                regra_distribuicao TEXT
            )
        """))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS epi_itens (
                codigo TEXT PRIMARY KEY,
                descricao TEXT,
                saldo NUMERIC DEFAULT 0,
                umb TEXT,
                nome_conta TEXT,
                codigo_conta TEXT,
                familia TEXT,
                deposito TEXT,
                custo_unitario NUMERIC DEFAULT 0
            )
        """))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS epi_solicitacoes (
                id TEXT PRIMARY KEY,
                numero TEXT,
                solicitante TEXT NOT NULL,
                setor_id TEXT,
                setor_nome TEXT,
                regra_distribuicao TEXT,
                motivo TEXT NOT NULL,
                data_hora TEXT NOT NULL,
                status TEXT NOT NULL,
                itens_json TEXT NOT NULL
            )
        """))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS epi_config (
                key TEXT PRIMARY KEY,
                value TEXT
            )
        """))


@app.on_event("startup")
def on_startup():
    init_db()


# ---------- helpers ----------
def parse_qtd_emb(s):
    if not s:
        return 1
    m = re.search(r"([\d.,]+)", str(s))
    return float(m.group(1).replace(",", ".")) if m else 1


def um_from_emb(s):
    m = re.search(r"[A-Za-zÀ-ÿ]+", str(s or ""))
    return m.group(0) if m else "UN"


def slugify(s):
    s = s.lower()
    s = re.sub(r"[^a-z0-9]+", "_", s).strip("_")
    return s[:24]


def parse_leading_number(s):
    m = re.search(r"([\d.,]+)", str(s or ""))
    return float(m.group(1).replace(",", ".")) if m else 0


def gen_id(prefix):
    return f"{prefix}_{int(datetime.now().timestamp()*1000)}_{''.join(random.choices(string.digits, k=3))}"


def inicio_janela_atual(janela):
    """Retorna o datetime (UTC) do início da janela de pedidos mais recente, com base na configuração."""
    now = datetime.now(timezone.utc)
    for i in range(8):
        d = now - timedelta(days=i)
        candidato = d.replace(hour=janela["inicio"]["hora"], minute=janela["inicio"]["minuto"], second=0, microsecond=0)
        # weekday(): Monday=0..Sunday=6 -> convertendo pro padrão dia (Domingo=0..Sábado=6) usado no frontend
        dia_padrao = (candidato.weekday() + 1) % 7
        if dia_padrao == janela["inicio"]["dia"] and candidato <= now:
            return candidato
    return None


# ---------- schemas ----------
class ItemPedido(BaseModel):
    codigo: str
    qtd: float


class NovoPedido(BaseModel):
    lojaId: str
    itens: list[ItemPedido]


class NovoSocio(BaseModel):
    nome: str


class NovaLoja(BaseModel):
    nome: str
    cnpj: Optional[str] = ""
    cidade: Optional[str] = ""
    socioId: str


class NovoItemEstoque(BaseModel):
    socioId: str
    codigo: str
    descricao: str
    und: Optional[str] = "UN"
    qtdEmb: Optional[str] = "1 UN"
    grupo: Optional[str] = "GERAL"


class SaldoUpdate(BaseModel):
    saldo: float


class StatusUpdate(BaseModel):
    status: str


class Janela(BaseModel):
    inicio: dict
    fim: dict


class TesteAberto(BaseModel):
    ligado: bool


# ---------- leitura ----------
@app.get("/api/data")
def get_all_data():
    with engine.begin() as conn:
        socios = [dict(r._mapping) for r in conn.execute(text("SELECT id, nome FROM socios ORDER BY nome"))]
        lojas = [dict(r._mapping) for r in conn.execute(text(
            "SELECT id, nome, cnpj, cidade, socio_id as \"socioId\" FROM lojas ORDER BY nome"))]

        estoques = {}
        for r in conn.execute(text(
                "SELECT socio_id, codigo, descricao, und, qtd_emb as \"qtdEmb\", grupo, saldo, "
                "valor_unitario as \"valorUnitario\" FROM estoque")):
            row = dict(r._mapping)
            sid = row.pop("socio_id")
            row["saldo"] = float(row["saldo"])
            row["valorUnitario"] = float(row["valorUnitario"]) if row["valorUnitario"] is not None else 0
            estoques.setdefault(sid, []).append(row)

        pedidos = []
        for r in conn.execute(text(
                "SELECT id, numero, loja_id as \"lojaId\", loja_nome as \"lojaNome\", "
                "socio_id as \"socioId\", data_hora as \"dataHora\", status, itens_json "
                "FROM pedidos ORDER BY data_hora DESC")):
            row = dict(r._mapping)
            row["itens"] = json.loads(row.pop("itens_json"))
            pedidos.append(row)

        janela = {"inicio": {"dia": 5, "hora": 18, "minuto": 0}, "fim": {"dia": 0, "hora": 23, "minuto": 59}}
        teste_aberto = False
        for r in conn.execute(text("SELECT key, value FROM config")):
            row = dict(r._mapping)
            if row["key"] == "janela":
                janela = json.loads(row["value"])
            if row["key"] == "testeAberto":
                teste_aberto = row["value"] == "true"

    return {
        "socios": socios, "lojas": lojas, "estoques": estoques,
        "pedidos": pedidos, "janela": janela, "testeAberto": teste_aberto,
    }


# ---------- criar pedido (transação real, com trava de linha) ----------
@app.post("/api/pedidos")
def criar_pedido(payload: NovoPedido):
    with engine.begin() as conn:  # transação: tudo ou nada
        loja = conn.execute(text(
            "SELECT id, nome, socio_id FROM lojas WHERE id = :id"), {"id": payload.lojaId}).fetchone()
        if not loja:
            raise HTTPException(404, "Loja não encontrada.")
        socio_id = loja.socio_id

        teste_row = conn.execute(text("SELECT value FROM config WHERE key='testeAberto'")).fetchone()
        teste_aberto = teste_row and teste_row.value == "true"
        if not teste_aberto:
            janela_row = conn.execute(text("SELECT value FROM config WHERE key='janela'")).fetchone()
            janela = json.loads(janela_row.value) if janela_row else {"inicio": {"dia": 5, "hora": 18, "minuto": 0}}
            inicio = inicio_janela_atual(janela)
            if inicio:
                existente = conn.execute(text(
                    "SELECT id, numero FROM pedidos WHERE loja_id = :lid AND data_hora >= :inicio ORDER BY data_hora DESC LIMIT 1"
                ), {"lid": payload.lojaId, "inicio": inicio.isoformat()}).fetchone()
                if existente:
                    raise HTTPException(409, f"Essa loja já tem um pedido nesta janela ({existente.numero}). Edite o pedido existente em vez de criar um novo.")

        itens_resp = []
        for item in payload.itens:
            if not item.qtd or item.qtd <= 0:
                continue
            # SELECT ... FOR UPDATE trava a linha até o fim da transação — concorrência real
            row = conn.execute(text(
                "SELECT descricao, qtd_emb, grupo, saldo FROM estoque "
                "WHERE socio_id = :sid AND codigo = :cod FOR UPDATE"
            ), {"sid": socio_id, "cod": item.codigo}).fetchone()
            if not row:
                continue
            saldo_antes = float(row.saldo)
            qtd_total = item.qtd * parse_qtd_emb(row.qtd_emb)
            if qtd_total > saldo_antes:
                raise HTTPException(400, f"O item {item.codigo} ({row.descricao}) excede o saldo disponível ({saldo_antes}). Reduza a quantidade.")
            saldo_depois = saldo_antes - qtd_total
            conn.execute(text(
                "UPDATE estoque SET saldo = :novo WHERE socio_id = :sid AND codigo = :cod"
            ), {"novo": saldo_depois, "sid": socio_id, "cod": item.codigo})
            itens_resp.append({
                "codigo": item.codigo, "descricao": row.descricao, "grupo": row.grupo,
                "qtdSolicitada": item.qtd, "qtdTotal": f"{qtd_total} {um_from_emb(row.qtd_emb)}",
                "saldoAntes": saldo_antes, "saldoDepois": saldo_depois,
                "saldoInsuficiente": saldo_depois < 0,
            })

        if not itens_resp:
            raise HTTPException(400, "Nenhum item válido selecionado.")

        now = datetime.now(timezone.utc)
        numero = "PS-" + now.strftime("%y%m%d") + "-" + "".join(random.choices(string.digits, k=3))
        pedido_id = gen_id("pedido")
        pedido = {
            "id": pedido_id, "numero": numero, "lojaId": loja.id, "lojaNome": loja.nome,
            "socioId": socio_id, "dataHora": now.isoformat(), "status": "Aguardando Rota",
            "itens": itens_resp,
        }
        conn.execute(text(
            "INSERT INTO pedidos (id, numero, loja_id, loja_nome, socio_id, data_hora, status, itens_json) "
            "VALUES (:id, :numero, :loja_id, :loja_nome, :socio_id, :data_hora, :status, :itens_json)"
        ), {
            "id": pedido_id, "numero": numero, "loja_id": loja.id, "loja_nome": loja.nome,
            "socio_id": socio_id, "data_hora": now.isoformat(), "status": "Aguardando Rota",
            "itens_json": json.dumps(itens_resp),
        })
    return pedido


# ---------- editar pedido existente (só permitido com a janela aberta, verificado no frontend) ----------
@app.put("/api/pedidos/{pedido_id}")
def editar_pedido(pedido_id: str, payload: NovoPedido):
    with engine.begin() as conn:
        pedido = conn.execute(text(
            "SELECT * FROM pedidos WHERE id = :id"), {"id": pedido_id}).fetchone()
        if not pedido:
            raise HTTPException(404, "Pedido não encontrado.")
        socio_id = pedido.socio_id
        itens_antigos = json.loads(pedido.itens_json)

        # reverte a baixa de estoque dos itens antigos
        for old in itens_antigos:
            row = conn.execute(text(
                "SELECT saldo FROM estoque WHERE socio_id = :sid AND codigo = :cod FOR UPDATE"
            ), {"sid": socio_id, "cod": old["codigo"]}).fetchone()
            if row:
                qtd_total_antiga = parse_leading_number(old.get("qtdTotal"))
                conn.execute(text(
                    "UPDATE estoque SET saldo = :novo WHERE socio_id = :sid AND codigo = :cod"
                ), {"novo": float(row.saldo) + qtd_total_antiga, "sid": socio_id, "cod": old["codigo"]})

        # aplica a baixa dos itens novos
        itens_resp = []
        for item in payload.itens:
            if not item.qtd or item.qtd <= 0:
                continue
            row = conn.execute(text(
                "SELECT descricao, qtd_emb, grupo, saldo FROM estoque "
                "WHERE socio_id = :sid AND codigo = :cod FOR UPDATE"
            ), {"sid": socio_id, "cod": item.codigo}).fetchone()
            if not row:
                continue
            saldo_antes = float(row.saldo)
            qtd_total = item.qtd * parse_qtd_emb(row.qtd_emb)
            if qtd_total > saldo_antes:
                raise HTTPException(400, f"O item {item.codigo} ({row.descricao}) excede o saldo disponível ({saldo_antes}). Reduza a quantidade.")
            saldo_depois = saldo_antes - qtd_total
            conn.execute(text(
                "UPDATE estoque SET saldo = :novo WHERE socio_id = :sid AND codigo = :cod"
            ), {"novo": saldo_depois, "sid": socio_id, "cod": item.codigo})
            itens_resp.append({
                "codigo": item.codigo, "descricao": row.descricao, "grupo": row.grupo,
                "qtdSolicitada": item.qtd, "qtdTotal": f"{qtd_total} {um_from_emb(row.qtd_emb)}",
                "saldoAntes": saldo_antes, "saldoDepois": saldo_depois,
                "saldoInsuficiente": saldo_depois < 0,
            })

        if not itens_resp:
            raise HTTPException(400, "Nenhum item válido selecionado.")

        conn.execute(text(
            "UPDATE pedidos SET itens_json = :itens WHERE id = :id"
        ), {"itens": json.dumps(itens_resp), "id": pedido_id})

    return {
        "id": pedido.id, "numero": pedido.numero, "lojaId": pedido.loja_id, "lojaNome": pedido.loja_nome,
        "socioId": socio_id, "dataHora": pedido.data_hora, "status": pedido.status, "itens": itens_resp,
    }


# ---------- demais ações ----------
@app.patch("/api/pedidos/{pedido_id}/status")
def atualizar_status(pedido_id: str, body: StatusUpdate):
    with engine.begin() as conn:
        result = conn.execute(text(
            "UPDATE pedidos SET status = :status WHERE id = :id"
        ), {"status": body.status, "id": pedido_id})
        if result.rowcount == 0:
            raise HTTPException(404, "Pedido não encontrado.")
    return {"ok": True}


@app.post("/api/socios")
def adicionar_socio(body: NovoSocio):
    new_id = f"socio_{slugify(body.nome)}_{random.randint(100,999)}"
    with engine.begin() as conn:
        conn.execute(text("INSERT INTO socios (id, nome) VALUES (:id, :nome)"), {"id": new_id, "nome": body.nome})
    return {"id": new_id, "nome": body.nome}


@app.post("/api/lojas")
def salvar_loja(body: NovaLoja):
    new_id = f"loja_{slugify(body.nome)}_{random.randint(100,999)}"
    with engine.begin() as conn:
        conn.execute(text(
            "INSERT INTO lojas (id, nome, cnpj, cidade, socio_id) VALUES (:id,:nome,:cnpj,:cidade,:socio_id)"
        ), {"id": new_id, "nome": body.nome, "cnpj": body.cnpj, "cidade": body.cidade, "socio_id": body.socioId})
    return {"id": new_id, **body.dict()}


@app.post("/api/estoque")
def salvar_item_estoque(body: NovoItemEstoque):
    with engine.begin() as conn:
        conn.execute(text(
            "INSERT INTO estoque (socio_id, codigo, descricao, und, qtd_emb, grupo, saldo) "
            "VALUES (:sid,:cod,:desc,:und,:emb,:grp,0) "
            "ON CONFLICT (socio_id, codigo) DO UPDATE SET descricao=:desc, und=:und, qtd_emb=:emb, grupo=:grp"
        ), {"sid": body.socioId, "cod": body.codigo, "desc": body.descricao,
            "und": body.und, "emb": body.qtdEmb, "grp": body.grupo})
    return {"ok": True}


@app.delete("/api/estoque/{socio_id}/{codigo}")
def remover_item_estoque(socio_id: str, codigo: str):
    with engine.begin() as conn:
        conn.execute(text(
            "DELETE FROM estoque WHERE socio_id = :sid AND codigo = :cod"
        ), {"sid": socio_id, "cod": codigo})
    return {"ok": True}


@app.patch("/api/estoque/{socio_id}/{codigo}")
def atualizar_saldo(socio_id: str, codigo: str, body: SaldoUpdate):
    with engine.begin() as conn:
        conn.execute(text(
            "UPDATE estoque SET saldo = :saldo WHERE socio_id = :sid AND codigo = :cod"
        ), {"saldo": body.saldo, "sid": socio_id, "cod": codigo})
    return {"ok": True}


@app.put("/api/config/janela")
def salvar_janela(body: Janela):
    with engine.begin() as conn:
        conn.execute(text(
            "INSERT INTO config (key, value) VALUES ('janela', :v) "
            "ON CONFLICT (key) DO UPDATE SET value = :v"
        ), {"v": json.dumps(body.dict())})
    return {"ok": True}


@app.put("/api/config/teste")
def salvar_teste_aberto(body: TesteAberto):
    with engine.begin() as conn:
        conn.execute(text(
            "INSERT INTO config (key, value) VALUES ('testeAberto', :v) "
            "ON CONFLICT (key) DO UPDATE SET value = :v"
        ), {"v": "true" if body.ligado else "false"})
    return {"ok": True}


@app.post("/api/estoque/importar")
async def importar_estoque(file: UploadFile):
    import openpyxl
    from io import BytesIO

    conteudo = await file.read()
    try:
        wb = openpyxl.load_workbook(BytesIO(conteudo), data_only=True)
    except Exception:
        raise HTTPException(400, "Não consegui abrir o arquivo. Confirme que é um .xlsx válido.")

    resultado = {"abas_processadas": [], "abas_ignoradas": [], "itens_importados": 0}

    with engine.begin() as conn:
        socios_por_codigo = {}
        for r in conn.execute(text("SELECT id, codigo FROM socios WHERE codigo IS NOT NULL")):
            socios_por_codigo[r.codigo.strip().upper()] = r.id

        for sheet_name in wb.sheetnames:
            codigo_aba = sheet_name.strip().upper()
            socio_id = socios_por_codigo.get(codigo_aba)
            if not socio_id:
                resultado["abas_ignoradas"].append(sheet_name)
                continue

            ws = wb[sheet_name]
            rows = list(ws.iter_rows(min_row=2, values_only=True))  # pula cabeçalho

            # substitui por completo o estoque desse sócio pelo que veio no arquivo
            conn.execute(text("DELETE FROM estoque WHERE socio_id = :sid"), {"sid": socio_id})

            for row in rows:
                if not row or row[0] is None:
                    continue
                codigo_item = str(row[0]).strip()
                descricao = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                saldo = row[2] if len(row) > 2 and row[2] is not None else 0
                valor_acumulado = row[3] if len(row) > 3 and row[3] is not None else 0
                valor_unitario = (valor_acumulado / saldo) if saldo else 0
                conn.execute(text(
                    "INSERT INTO estoque (socio_id, codigo, descricao, und, qtd_emb, grupo, saldo, valor_acumulado, valor_unitario) "
                    "VALUES (:sid,:cod,:descricao,'UN','1 UN','GERAL',:saldo,:valor_acum,:valor_unit)"
                ), {"sid": socio_id, "cod": codigo_item, "descricao": descricao, "saldo": saldo,
                    "valor_acum": valor_acumulado, "valor_unit": valor_unitario})
                resultado["itens_importados"] += 1

            resultado["abas_processadas"].append(sheet_name)

    return resultado


@app.get("/")
def root():
    return {"status": "ok", "service": "Frosty - Pedidos de Secos API"}


@app.get("/api/seed-inicial")
def seed_inicial():
    from seed_data import SOCIOS_SEED, LOJAS_SEED, ESTOQUES_SEED
    with engine.begin() as conn:
        ja_tem = conn.execute(text("SELECT COUNT(*) FROM socios")).scalar()
        if ja_tem and ja_tem > 0:
            return {"ok": False, "mensagem": "Já existem dados — nada foi alterado. Se quiser popular de novo, apague as tabelas antes."}

        for s in SOCIOS_SEED:
            conn.execute(text("INSERT INTO socios (id, nome) VALUES (:id,:nome)"), s)
        for l in LOJAS_SEED:
            conn.execute(text(
                "INSERT INTO lojas (id, nome, cnpj, cidade, socio_id) VALUES (:id,:nome,:cnpj,:cidade,:socioId)"
            ), l)
        total_itens = 0
        for socio_id, itens in ESTOQUES_SEED.items():
            for it in itens:
                conn.execute(text(
                    "INSERT INTO estoque (socio_id, codigo, descricao, und, qtd_emb, grupo, saldo) "
                    "VALUES (:sid,:codigo,:descricao,:und,:qtdEmb,:grupo,:saldo)"
                ), {"sid": socio_id, **it})
                total_itens += 1
        conn.execute(text(
            "INSERT INTO config (key, value) VALUES ('janela', :v)"
        ), {"v": json.dumps({"inicio": {"dia": 5, "hora": 18, "minuto": 0}, "fim": {"dia": 0, "hora": 23, "minuto": 59}})})
        conn.execute(text("INSERT INTO config (key, value) VALUES ('testeAberto', 'false')"))

    return {"ok": True, "socios": len(SOCIOS_SEED), "lojas": len(LOJAS_SEED), "itens_estoque": total_itens}


# ============================================================
#  EPI / MATERIAL DE EXPEDIENTE
# ============================================================

class ItemEpiPedido(BaseModel):
    codigo: Optional[str] = None   # vazio quando motivo = teste (item fora do catálogo)
    descricao: Optional[str] = None
    qtd: float


class NovaSolicitacaoEpi(BaseModel):
    solicitante: str
    setorId: str
    motivo: str  # "Aquisição" | "Troca" | "Teste"
    itens: list[ItemEpiPedido]


class StatusEpiUpdate(BaseModel):
    status: str


class JanelaEpi(BaseModel):
    inicio: dict
    fim: dict


@app.get("/api/epi/data")
def epi_get_all_data():
    with engine.begin() as conn:
        setores = [dict(r._mapping) for r in conn.execute(text(
            "SELECT id, nome, regra_distribuicao as \"regraDistribuicao\" FROM epi_setores ORDER BY nome"))]
        itens = []
        for r in conn.execute(text(
                "SELECT codigo, descricao, saldo, umb, nome_conta as \"nomeConta\", codigo_conta as \"codigoConta\", "
                "familia, deposito, custo_unitario as \"custoUnitario\" FROM epi_itens ORDER BY familia, descricao")):
            row = dict(r._mapping)
            row["saldo"] = float(row["saldo"]) if row["saldo"] is not None else 0
            itens.append(row)

        solicitacoes = []
        for r in conn.execute(text(
                "SELECT id, numero, solicitante, setor_id as \"setorId\", setor_nome as \"setorNome\", "
                "regra_distribuicao as \"regraDistribuicao\", motivo, data_hora as \"dataHora\", status, itens_json "
                "FROM epi_solicitacoes ORDER BY data_hora DESC")):
            row = dict(r._mapping)
            row["itens"] = json.loads(row.pop("itens_json"))
            solicitacoes.append(row)

        janela = {"inicio": {"dia": 1, "hora": 8, "minuto": 0}, "fim": {"dia": 3, "hora": 18, "minuto": 0}}
        for r in conn.execute(text("SELECT value FROM epi_config WHERE key='janela'")):
            janela = json.loads(r.value)

    return {"setores": setores, "itens": itens, "solicitacoes": solicitacoes, "janela": janela}


@app.post("/api/epi/solicitacoes")
def epi_criar_solicitacao(payload: NovaSolicitacaoEpi):
    with engine.begin() as conn:
        setor = conn.execute(text(
            "SELECT id, nome, regra_distribuicao FROM epi_setores WHERE id = :id"), {"id": payload.setorId}).fetchone()
        if not setor:
            raise HTTPException(404, "Setor não encontrado.")

        itens_resp = []
        eh_teste = payload.motivo.strip().lower() == "teste"

        for item in payload.itens:
            if not item.qtd or item.qtd <= 0:
                continue
            if eh_teste:
                # motivo teste: item livre, sem catálogo nem controle de estoque
                itens_resp.append({
                    "codigo": item.codigo or "", "descricao": item.descricao or "(item de teste)",
                    "qtd": item.qtd, "codigoConta": "", "nomeConta": "",
                })
            else:
                row = conn.execute(text(
                    "SELECT descricao, saldo, codigo_conta, nome_conta FROM epi_itens WHERE codigo = :cod FOR UPDATE"
                ), {"cod": item.codigo}).fetchone()
                if not row:
                    continue
                saldo_antes = float(row.saldo)
                if item.qtd > saldo_antes:
                    raise HTTPException(400, f"O item {item.codigo} ({row.descricao}) excede o saldo disponível ({saldo_antes}). Reduza a quantidade.")
                saldo_depois = saldo_antes - item.qtd
                conn.execute(text(
                    "UPDATE epi_itens SET saldo = :novo WHERE codigo = :cod"
                ), {"novo": saldo_depois, "cod": item.codigo})
                itens_resp.append({
                    "codigo": item.codigo, "descricao": row.descricao, "qtd": item.qtd,
                    "codigoConta": row.codigo_conta, "nomeConta": row.nome_conta,
                })

        if not itens_resp:
            raise HTTPException(400, "Nenhum item válido selecionado.")

        now = datetime.now(timezone.utc)
        numero = "EPI-" + now.strftime("%y%m%d") + "-" + "".join(random.choices(string.digits, k=3))
        sol_id = gen_id("episol")
        solicitacao = {
            "id": sol_id, "numero": numero, "solicitante": payload.solicitante,
            "setorId": setor.id, "setorNome": setor.nome, "regraDistribuicao": setor.regra_distribuicao,
            "motivo": payload.motivo, "dataHora": now.isoformat(), "status": "Aguardando Separação",
            "itens": itens_resp,
        }
        conn.execute(text(
            "INSERT INTO epi_solicitacoes (id, numero, solicitante, setor_id, setor_nome, regra_distribuicao, "
            "motivo, data_hora, status, itens_json) VALUES "
            "(:id,:numero,:solicitante,:setor_id,:setor_nome,:regra,:motivo,:data_hora,:status,:itens_json)"
        ), {
            "id": sol_id, "numero": numero, "solicitante": payload.solicitante, "setor_id": setor.id,
            "setor_nome": setor.nome, "regra": setor.regra_distribuicao, "motivo": payload.motivo,
            "data_hora": now.isoformat(), "status": "Aguardando Separação", "itens_json": json.dumps(itens_resp),
        })

    return solicitacao


@app.patch("/api/epi/solicitacoes/{sol_id}/status")
def epi_atualizar_status(sol_id: str, body: StatusEpiUpdate):
    with engine.begin() as conn:
        result = conn.execute(text(
            "UPDATE epi_solicitacoes SET status = :status WHERE id = :id"
        ), {"status": body.status, "id": sol_id})
        if result.rowcount == 0:
            raise HTTPException(404, "Solicitação não encontrada.")
    return {"ok": True}


@app.put("/api/epi/config/janela")
def epi_salvar_janela(body: JanelaEpi):
    with engine.begin() as conn:
        ok = conn.execute(text(
            "UPDATE epi_config SET value = :v WHERE key = 'janela'"
        ), {"v": json.dumps(body.dict())})
        if ok.rowcount == 0:
            conn.execute(text("INSERT INTO epi_config (key, value) VALUES ('janela', :v)"), {"v": json.dumps(body.dict())})
    return {"ok": True}


@app.get("/api/epi/solicitacoes/{sol_id}/pdf")
def epi_gerar_pdf(sol_id: str):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle
    from reportlab.pdfgen import canvas
    from io import BytesIO
    from fastapi.responses import StreamingResponse

    with engine.begin() as conn:
        row = conn.execute(text("SELECT * FROM epi_solicitacoes WHERE id = :id"), {"id": sol_id}).fetchone()
    if not row:
        raise HTTPException(404, "Solicitação não encontrada.")
    itens = json.loads(row.itens_json)
    dt = datetime.fromisoformat(row.data_hora)

    AZUL = colors.HexColor("#1B2FE0")
    VERDE = colors.HexColor("#3FC340")
    CINZA_HEADER = colors.HexColor("#D9E1F2")

    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    largura, altura = A4
    margem = 1.6*cm

    # ---- cabeçalho azul com marca "frosty" ----
    banner_h = 1.7*cm
    c.setFillColor(AZUL)
    c.rect(0, altura-banner_h, largura, banner_h, fill=1, stroke=0)
    c.setFillColor(VERDE)
    c.roundRect(margem, altura-banner_h+0.3*cm, 2.8*cm, 1.1*cm, 8, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 13)
    c.drawCentredString(margem+1.4*cm, altura-banner_h+0.65*cm, "frosty")
    c.setFont("Helvetica-Bold", 17)
    c.drawCentredString(largura/2, altura-banner_h+0.62*cm, "ATENDIMENTO DE RESERVAS")

    # ---- dados do cabeçalho ----
    y = altura - banner_h - 0.9*cm
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(margem, y, "SETOR:")
    c.setFont("Helvetica", 10)
    c.drawString(margem+1.8*cm, y, row.setor_nome or "")
    c.setFont("Helvetica-Bold", 10)
    c.drawString(largura/2, y, "Solicitante:")
    c.setFont("Helvetica", 10)
    c.drawString(largura/2+2.4*cm, y, row.solicitante)

    y -= 0.55*cm
    c.setFont("Helvetica-Bold", 10)
    c.drawString(margem, y, "CC:")
    c.setFont("Helvetica", 10)
    c.drawString(margem+1.8*cm, y, str(row.regra_distribuicao or "-"))
    c.setFont("Helvetica-Bold", 10)
    c.drawString(largura/2, y, "Nº solicitação:")
    c.setFont("Helvetica", 10)
    c.drawString(largura/2+2.8*cm, y, row.numero)

    y -= 1*cm

    # ---- tabela de itens ----
    data = [["ITEM", "CÓDIGO", "DESCRIÇÃO DO MATERIAL", "UMB", "QNTD", "COD CONTA", "ATENDIDA"]]
    for i, it in enumerate(itens, start=1):
        data.append([
            str(i), it.get("codigo") or "-", (it.get("descricao") or "")[:55],
            it.get("umb") or "-", str(it.get("qtd")), it.get("codigoConta") or "-", "S (   )   N (   )"
        ])

    col_widths = [1.1*cm, 2.4*cm, 7.3*cm, 1.4*cm, 1.4*cm, 2.6*cm, 3.0*cm]
    tbl = Table(data, colWidths=col_widths)
    tbl.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), CINZA_HEADER),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 8.5),
        ('GRID', (0,0), (-1,-1), 0.6, AZUL),
        ('ALIGN', (0,0), (0,-1), 'CENTER'),
        ('ALIGN', (3,0), (4,-1), 'CENTER'),
        ('ALIGN', (6,0), (6,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    tw, th = tbl.wrapOn(c, largura-2*margem, altura)
    if th > y - 4*cm:  # não cabe tudo na página: começa de novo com margem cheia (pedidos muito grandes)
        c.showPage()
        y = altura - margem
    tbl.drawOn(c, margem, y-th)
    y = y - th - 1.1*cm

    # ---- rodapé: assinatura e motivo ----
    if y < 3*cm:
        c.showPage(); y = altura - margem
    c.setFont("Helvetica", 10)
    c.drawString(margem, y, "Entregue por: ______________________________________")
    c.drawString(margem+10.5*cm, y, f"Motivo: {row.motivo}")
    y -= 0.7*cm
    c.drawString(margem, y, f"Data da solicitação: {dt.strftime('%d/%m/%Y')}")

    c.showPage()
    c.save()
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf", headers={
        "Content-Disposition": f'inline; filename="solicitacao_{row.numero}.pdf"'
    })


@app.get("/api/epi/seed-inicial")
def epi_seed_inicial():
    from epi_seed_data import EPI_SETORES_SEED, EPI_ITENS_SEED
    with engine.begin() as conn:
        ja_tem = conn.execute(text("SELECT COUNT(*) FROM epi_setores")).scalar()
        if ja_tem and ja_tem > 0:
            return {"ok": False, "mensagem": "Já existem dados de EPI — nada foi alterado."}

        for i, s in enumerate(EPI_SETORES_SEED):
            conn.execute(text(
                "INSERT INTO epi_setores (id, nome, regra_distribuicao) VALUES (:id,:nome,:regra)"
            ), {"id": f"setor_{i}", "nome": s["nome"], "regra": s["regraDistribuicao"]})

        for it in EPI_ITENS_SEED:
            conn.execute(text(
                "INSERT INTO epi_itens (codigo, descricao, saldo, umb, nome_conta, codigo_conta, familia, deposito, custo_unitario) "
                "VALUES (:codigo,:descricao,:saldo,:umb,:nomeConta,:codigoConta,:familia,:deposito,:custoUnitario) "
                "ON CONFLICT (codigo) DO NOTHING"
            ), it)

        conn.execute(text(
            "INSERT INTO epi_config (key, value) VALUES ('janela', :v) ON CONFLICT (key) DO NOTHING"
        ), {"v": json.dumps({"inicio": {"dia": 1, "hora": 8, "minuto": 0}, "fim": {"dia": 3, "hora": 18, "minuto": 0}})})

    return {"ok": True, "setores": len(EPI_SETORES_SEED), "itens": len(EPI_ITENS_SEED)}
